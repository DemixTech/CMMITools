#!/usr/bin/env python3
"""
Audit CAS Plan - Validates _FieldMap against actual sheet content.

This script reads the _FieldMap sheet and verifies that:
1. The field exists at the specified row
2. The field name matches Column A in the target sheet
3. The cell type (value/formula) matches the actual cell

Results are written to Column G (Notes) with "ok" or details of mismatch.

Author: Claude AI Assistant
Version: 1.0.0
"""

import argparse
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    exit(1)

# Styling
OK_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
OK_FONT = Font(color='006100')
ERROR_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
ERROR_FONT = Font(color='9C0006')
WARNING_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
WARNING_FONT = Font(color='9C5700')

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class CASPlanAuditor:
    """Audits _FieldMap against actual sheet content."""
    
    def __init__(self, file_path: str, dry_run: bool = False):
        self.file_path = Path(file_path)
        self.dry_run = dry_run
        self.results = {
            'ok': 0,
            'mismatch': 0,
            'warning': 0,
            'errors': []
        }
    
    def validate_file(self) -> bool:
        """Validate that the file exists."""
        if not self.file_path.exists():
            logger.error(f"File not found: {self.file_path}")
            return False
        return True
    
    def is_formula(self, cell) -> bool:
        """Check if a cell contains a formula."""
        if cell.value is None:
            return False
        if isinstance(cell.value, str) and str(cell.value).startswith('='):
            return True
        return False
    
    def normalize(self, value) -> str:
        """Normalize string for comparison."""
        if value is None:
            return ""
        return str(value).strip().lower()
    
    def audit_field(self, wb, sheet_name: str, heading: str, field_name: str, 
                    expected_row: int, expected_type: str) -> Dict[str, Any]:
        """
        Audit a single field entry.
        
        Returns:
            Dictionary with audit result
        """
        result = {
            'status': 'ok',
            'message': 'ok',
            'details': {}
        }
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            result['status'] = 'error'
            result['message'] = f"Sheet '{sheet_name}' not found"
            return result
        
        sheet = wb[sheet_name]
        
        # Get actual cell values
        actual_cell_a = sheet.cell(row=expected_row, column=1)
        actual_cell_b = sheet.cell(row=expected_row, column=2)
        
        actual_field_name = str(actual_cell_a.value).strip() if actual_cell_a.value else ""
        actual_type = "formula" if self.is_formula(actual_cell_b) else "value"
        
        # For headings, Column B should be empty
        if expected_type == "heading":
            actual_type = "heading"
        
        issues = []
        
        # Check 1: Field name matches (for non-heading rows)
        if field_name != "[HEADING]":
            if self.normalize(actual_field_name) != self.normalize(field_name):
                issues.append(f"Name mismatch: expected '{field_name}', found '{actual_field_name}'")
                result['details']['expected_name'] = field_name
                result['details']['actual_name'] = actual_field_name
        else:
            # For headings, check the heading name matches
            if self.normalize(actual_field_name) != self.normalize(heading):
                issues.append(f"Heading mismatch: expected '{heading}', found '{actual_field_name}'")
                result['details']['expected_heading'] = heading
                result['details']['actual_heading'] = actual_field_name
        
        # Check 2: Type matches (only for non-heading fields)
        if expected_type != "heading":
            if actual_type != expected_type:
                issues.append(f"Type mismatch: expected '{expected_type}', found '{actual_type}'")
                result['details']['expected_type'] = expected_type
                result['details']['actual_type'] = actual_type
        
        # Determine final status
        if issues:
            result['status'] = 'mismatch'
            result['message'] = '; '.join(issues)
        
        return result
    
    def run(self) -> Dict[str, Any]:
        """Execute the audit."""
        logger.info("=" * 60)
        logger.info("CAS Plan Audit - Starting (v1.0.0)")
        logger.info("=" * 60)
        logger.info(f"File: {self.file_path}")
        logger.info(f"Dry run: {self.dry_run}")
        
        if not self.validate_file():
            return {'success': False, 'error': 'File validation failed'}
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(self.file_path, keep_vba=True)
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return {'success': False, 'error': str(e)}
        
        # Check for _FieldMap sheet
        if '_FieldMap' not in wb.sheetnames:
            logger.error("_FieldMap sheet not found in workbook")
            wb.close()
            return {'success': False, 'error': '_FieldMap sheet not found'}
        
        field_map = wb['_FieldMap']
        
        logger.info(f"Found _FieldMap with {field_map.max_row - 1} entries")
        
        # Process each row in _FieldMap (skip header)
        for row in range(2, (field_map.max_row or 1) + 1):
            sheet_name = field_map.cell(row=row, column=1).value
            heading = field_map.cell(row=row, column=2).value or ""
            field_name = field_map.cell(row=row, column=3).value or ""
            expected_row = field_map.cell(row=row, column=4).value
            expected_type = field_map.cell(row=row, column=5).value or "value"
            
            if not sheet_name or not expected_row:
                continue
            
            # Audit this field
            result = self.audit_field(
                wb, 
                sheet_name, 
                heading, 
                field_name, 
                int(expected_row), 
                expected_type
            )
            
            # Write result to Column G (Notes)
            notes_cell = field_map.cell(row=row, column=7)
            
            if not self.dry_run:
                notes_cell.value = result['message']
                
                # Apply styling based on status
                if result['status'] == 'ok':
                    notes_cell.fill = OK_FILL
                    notes_cell.font = OK_FONT
                    self.results['ok'] += 1
                elif result['status'] == 'mismatch':
                    notes_cell.fill = ERROR_FILL
                    notes_cell.font = ERROR_FONT
                    self.results['mismatch'] += 1
                    self.results['errors'].append({
                        'row': row,
                        'sheet': sheet_name,
                        'field': field_name,
                        'message': result['message']
                    })
                else:
                    notes_cell.fill = WARNING_FILL
                    notes_cell.font = WARNING_FONT
                    self.results['warning'] += 1
            
            # Log mismatches
            if result['status'] != 'ok':
                logger.warning(f"Row {row}: {sheet_name}!{field_name} - {result['message']}")
        
        # Update header for Column G
        if not self.dry_run:
            header_cell = field_map.cell(row=1, column=7)
            header_cell.value = "Audit Result"
        
        # Save changes
        if not self.dry_run:
            try:
                wb.save(self.file_path)
                logger.info(f"Workbook saved with audit results")
            except Exception as e:
                logger.error(f"Failed to save workbook: {e}")
                wb.close()
                return {'success': False, 'error': f'Save failed: {e}'}
        
        wb.close()
        
        # Summary
        total = self.results['ok'] + self.results['mismatch'] + self.results['warning']
        
        logger.info("=" * 60)
        logger.info("CAS Plan Audit - Complete")
        logger.info(f"Total fields audited: {total}")
        logger.info(f"  OK: {self.results['ok']}")
        logger.info(f"  Mismatches: {self.results['mismatch']}")
        logger.info(f"  Warnings: {self.results['warning']}")
        logger.info("=" * 60)
        
        return {
            'success': True,
            'dry_run': self.dry_run,
            'total': total,
            'ok': self.results['ok'],
            'mismatch': self.results['mismatch'],
            'warning': self.results['warning'],
            'errors': self.results['errors'][:50]  # Limit for readability
        }
    
    def generate_report(self, results: Dict[str, Any], output_path: Optional[str] = None):
        """Generate a detailed report."""
        report_lines = [
            "=" * 70,
            "CAS PLAN AUDIT REPORT (v1.0.0)",
            "=" * 70,
            f"Generated: {datetime.now().isoformat()}",
            f"File: {self.file_path}",
            f"Dry Run: {results.get('dry_run', False)}",
            "",
            "SUMMARY",
            "-" * 50,
            f"Total fields audited: {results.get('total', 0)}",
            f"  OK (matches): {results.get('ok', 0)}",
            f"  Mismatches: {results.get('mismatch', 0)}",
            f"  Warnings: {results.get('warning', 0)}",
            "",
        ]
        
        if results.get('errors'):
            report_lines.extend([
                "MISMATCHES FOUND",
                "-" * 50,
            ])
            for err in results['errors']:
                report_lines.append(f"  Row {err['row']}: {err['sheet']}!{err['field']}")
                report_lines.append(f"    {err['message']}")
                report_lines.append("")
        else:
            report_lines.extend([
                "NO MISMATCHES FOUND",
                "-" * 50,
                "All fields in _FieldMap match the actual sheet content.",
                "",
            ])
        
        report_lines.extend([
            "COLUMN G LEGEND",
            "-" * 50,
            "  'ok' (green)  = Field matches _FieldMap definition",
            "  Error (red)   = Mismatch found (name or type)",
            "  Warning (yellow) = Other issue",
            "",
            "=" * 70,
            "END OF REPORT",
            "=" * 70,
        ])
        
        report = "\n".join(report_lines)
        
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(report)
            logger.info(f"Report saved to: {output_path}")
        
        return report


def main():
    parser = argparse.ArgumentParser(
        description='Audit CAS Plan _FieldMap against actual sheet content (v1.0.0)'
    )
    parser.add_argument(
        '--file', '-f',
        required=True,
        help='Path to the CAS Plan file containing _FieldMap'
    )
    parser.add_argument(
        '--dry-run', '-n',
        action='store_true',
        help='Preview audit without writing results'
    )
    parser.add_argument(
        '--report', '-r',
        help='Path to save detailed report (txt file)'
    )
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable verbose logging'
    )
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    auditor = CASPlanAuditor(
        file_path=args.file,
        dry_run=args.dry_run
    )
    
    results = auditor.run()
    
    if args.report:
        auditor.generate_report(results, args.report)
    else:
        print(auditor.generate_report(results))
    
    exit(0 if results.get('success') else 1)


if __name__ == '__main__':
    main()

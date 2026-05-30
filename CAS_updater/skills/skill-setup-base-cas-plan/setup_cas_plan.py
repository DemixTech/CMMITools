#!/usr/bin/env python3
"""
Setup Base CAS Plan - Populate CAS Plan workbook from startup/demix source file.

Handles three types of sheets:
1. Agreement 3, Planning, StartupInfo: Column B only, _FieldMap driven
2. Project&Support: Full row copy (cols A-AB), matched by WorkID
3. Staff: Copy columns A, C, D only; preserve formulas in B, E; 
   process only p# and s# WorkIDs (ignore o#); adjust row count

Version 3.1.0

Author: Claude AI Assistant
"""

import argparse
import logging
import shutil
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.drawing.image import Image
    from PIL import Image as PILImage
except ImportError:
    print("Error: openpyxl and Pillow are required.")
    print("Install with: pip install openpyxl Pillow")
    exit(1)

# Styling for source file marking
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
RED_FONT = Font(color='FF0000', bold=True)

# Image configuration
LOGO_ROW = 15
LOGO_CELL = 'B15'
LOGO_SHEET = 'StartupInfo'
LOGO_ROW_HEIGHT = 60
LOGO_MAX_WIDTH = 180
LOGO_MAX_HEIGHT = 75

# Sheet processing modes
FIELDMAP_SHEETS = ['Agreement 3', 'Planning', 'StartupInfo']
PROJECT_SUPPORT_SHEET = 'Project&Support'
STAFF_SHEET = 'Staff'

# Project&Support config
PS_DATA_START_ROW = 3
PS_MAX_COL = 28  # Column AB

# Staff config
STAFF_DATA_START_ROW = 3
STAFF_COPY_COLS = [1, 3, 4]  # A, C, D
STAFF_VALID_PREFIXES = ['p', 's']  # Only process p# and s# WorkIDs

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class CASPlanSetup:
    """Setup CAS Plan from source file."""
    
    def __init__(self, source_path: str, target_path: str, dry_run: bool = False):
        self.source_path = Path(source_path)
        self.target_path = Path(target_path)
        self.dry_run = dry_run
        self.changes: List[Dict[str, Any]] = []
        self.warnings: List[Dict[str, Any]] = []
        self.anomalies: List[Dict[str, Any]] = []
        self.preserved: List[Dict[str, Any]] = []
        self.field_map_entries: List[Dict[str, Any]] = []
        self.image_info: Optional[Dict[str, Any]] = None
        self.ps_stats = {'matched': 0, 'not_found': 0}
        self.staff_stats = {'rows_copied': 0, 'rows_deleted': 0, 'rows_skipped': 0}
    
    def validate_files(self) -> bool:
        if not self.source_path.exists():
            logger.error(f"Source file not found: {self.source_path}")
            return False
        if not self.target_path.exists():
            logger.error(f"Target file not found: {self.target_path}")
            return False
        return True
    
    def create_backup(self, file_path: Path) -> Optional[Path]:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = file_path.with_suffix(f'.backup_{timestamp}.xlsm')
        try:
            shutil.copy2(file_path, backup_path)
            logger.info(f"Backup created: {backup_path}")
            return backup_path
        except Exception as e:
            logger.error(f"Failed to create backup: {e}")
            return None
    
    def normalize(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip().lower()
    
    def is_formula(self, cell) -> bool:
        if cell.value is None:
            return False
        if isinstance(cell.value, str) and str(cell.value).startswith('='):
            return True
        return False
    
    def is_empty(self, value) -> bool:
        if value is None:
            return True
        if isinstance(value, str) and value.strip() == "":
            return True
        return False
    
    def is_valid_staff_workid(self, work_id) -> bool:
        """Check if WorkID is valid for Staff (p# or s# only)."""
        if not work_id:
            return False
        normalized = self.normalize(work_id)
        return any(normalized.startswith(prefix) for prefix in STAFF_VALID_PREFIXES)
    
    def extract_company_logo(self) -> Optional[Dict[str, Any]]:
        """Extract and resize company logo from source xlsx."""
        try:
            with zipfile.ZipFile(self.source_path, 'r') as zf:
                for img_name in zf.namelist():
                    if not img_name.startswith('xl/media/') or img_name == 'xl/media/':
                        continue
                    img_data = zf.read(img_name)
                    if len(img_data) == 0:
                        continue
                    try:
                        pil_img = PILImage.open(BytesIO(img_data))
                        w, h = pil_img.size
                        if w > 400 and h > 100 and w/h > 2:
                            logger.info(f"Found company logo: {img_name} ({w}x{h})")
                            aspect = w / h
                            if aspect > (LOGO_MAX_WIDTH / LOGO_MAX_HEIGHT):
                                new_w = LOGO_MAX_WIDTH
                                new_h = int(LOGO_MAX_WIDTH / aspect)
                            else:
                                new_h = LOGO_MAX_HEIGHT
                                new_w = int(LOGO_MAX_HEIGHT * aspect)
                            pil_img_resized = pil_img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)
                            resized_path = '/tmp/cas_plan_logo_resized.png'
                            pil_img_resized.save(resized_path, 'PNG')
                            return {'path': resized_path, 'orig_size': (w, h), 'new_size': (new_w, new_h)}
                    except Exception:
                        continue
        except Exception as e:
            logger.warning(f"Error extracting logo: {e}")
        return None
    
    def load_field_map(self, target_wb) -> bool:
        """Load _FieldMap from target workbook."""
        if '_FieldMap' not in target_wb.sheetnames:
            logger.error("_FieldMap sheet not found in target workbook")
            return False
        
        field_map = target_wb['_FieldMap']
        for row in range(2, (field_map.max_row or 1) + 1):
            sheet_name = field_map.cell(row=row, column=1).value
            heading = field_map.cell(row=row, column=2).value or ""
            field_name = field_map.cell(row=row, column=3).value or ""
            target_row = field_map.cell(row=row, column=4).value
            field_type = field_map.cell(row=row, column=5).value or "value"
            aliases = field_map.cell(row=row, column=6).value or ""
            
            if not sheet_name or not target_row:
                continue
            
            self.field_map_entries.append({
                'sheet': sheet_name,
                'heading': str(heading).strip(),
                'field_name': str(field_name).strip(),
                'target_row': int(target_row),
                'type': field_type,
                'aliases': [a.strip() for a in str(aliases).split(',') if a.strip()]
            })
        
        logger.info(f"Loaded {len(self.field_map_entries)} entries from _FieldMap")
        return True
    
    def build_source_index(self, source_wb, sheet_name: str, headings: List[str]) -> Dict[str, Tuple[int, Any]]:
        """Build index of source fields for FieldMap sheets."""
        if sheet_name not in source_wb.sheetnames:
            return {}
        sheet = source_wb[sheet_name]
        index = {}
        current_heading = ""
        normalized_headings = {self.normalize(h): h for h in headings}
        
        for row in range(1, (sheet.max_row or 0) + 1):
            cell_a = sheet.cell(row=row, column=1)
            cell_b = sheet.cell(row=row, column=2)
            field_name = str(cell_a.value).strip() if cell_a.value else ""
            if not field_name:
                continue
            normalized_field = self.normalize(field_name)
            if normalized_field in normalized_headings:
                current_heading = field_name
                key = f"{self.normalize(current_heading)}::[heading]"
                index[key] = (row, cell_b.value)
            else:
                key = f"{self.normalize(current_heading)}::{normalized_field}"
                index[key] = (row, cell_b.value)
        return index
    
    def get_headings_for_sheet(self, sheet_name: str) -> List[str]:
        return [e['heading'] for e in self.field_map_entries 
                if e['sheet'] == sheet_name and e['field_name'] == '[HEADING]']
    
    def find_source_field(self, source_index: Dict, heading: str, field_name: str, 
                          aliases: List[str]) -> Optional[Tuple[int, Any]]:
        key = f"{self.normalize(heading)}::{self.normalize(field_name)}"
        if key in source_index:
            return source_index[key]
        for alias in aliases:
            key = f"{self.normalize(heading)}::{self.normalize(alias)}"
            if key in source_index:
                return source_index[key]
        for index_key, value in source_index.items():
            if index_key.endswith(f"::{self.normalize(field_name)}"):
                return value
            for alias in aliases:
                if index_key.endswith(f"::{self.normalize(alias)}"):
                    return value
        return None
    
    def process_fieldmap_sheet(self, source_wb, target_wb, sheet_name: str) -> int:
        """Process a FieldMap-driven sheet."""
        sheet_entries = [e for e in self.field_map_entries if e['sheet'] == sheet_name]
        if not sheet_entries:
            return 0
        if sheet_name not in source_wb.sheetnames or sheet_name not in target_wb.sheetnames:
            return 0
        
        source_sheet = source_wb[sheet_name]
        target_sheet = target_wb[sheet_name]
        headings = self.get_headings_for_sheet(sheet_name)
        source_index = self.build_source_index(source_wb, sheet_name, headings)
        
        logger.info(f"Processing sheet '{sheet_name}' (FieldMap mode)")
        cells_copied = 0
        
        for entry in sheet_entries:
            if entry['field_name'] == '[HEADING]' or entry['type'] == 'formula':
                continue
            
            source_result = self.find_source_field(
                source_index, entry['heading'], entry['field_name'], entry['aliases']
            )
            if source_result is None:
                self.warnings.append({'sheet': sheet_name, 'field': entry['field_name'], 'reason': 'Not found in source'})
                continue
            
            source_row, source_value = source_result
            target_row = entry['target_row']
            target_cell = target_sheet.cell(row=target_row, column=2)
            
            if self.is_formula(target_cell):
                continue
            
            if self.is_empty(source_value) and not self.is_empty(target_cell.value):
                self.preserved.append({'sheet': sheet_name, 'field': entry['field_name'], 'value': str(target_cell.value)[:50]})
                continue
            
            if self.is_empty(source_value):
                continue
            
            is_anomaly = (source_row != target_row)
            if is_anomaly:
                self.anomalies.append({'sheet': sheet_name, 'field': entry['field_name'], 'source_row': source_row, 'target_row': target_row})
                if not self.dry_run:
                    source_sheet.cell(row=source_row, column=1).fill = YELLOW_FILL
                    source_sheet.cell(row=source_row, column=1).font = RED_FONT
                    source_sheet.cell(row=source_row, column=2).fill = YELLOW_FILL
                    source_sheet.cell(row=source_row, column=2).font = RED_FONT
            
            if not self.dry_run:
                target_cell.value = source_value
            
            cells_copied += 1
        
        return cells_copied
    
    def process_project_support(self, source_wb, target_wb) -> int:
        """Process Project&Support sheet (columns A to AB)."""
        if PROJECT_SUPPORT_SHEET not in source_wb.sheetnames:
            return 0
        if PROJECT_SUPPORT_SHEET not in target_wb.sheetnames:
            return 0
        
        source_sheet = source_wb[PROJECT_SUPPORT_SHEET]
        target_sheet = target_wb[PROJECT_SUPPORT_SHEET]
        
        logger.info(f"Processing sheet '{PROJECT_SUPPORT_SHEET}' (WorkID match, cols A-AB)")
        
        # Build target index
        target_index = {}
        for row in range(PS_DATA_START_ROW, target_sheet.max_row + 1):
            work_id = target_sheet.cell(row=row, column=1).value
            if work_id and str(work_id).strip():
                target_index[self.normalize(work_id)] = row
        
        cells_copied = 0
        
        for src_row in range(PS_DATA_START_ROW, source_sheet.max_row + 1):
            work_id = source_sheet.cell(row=src_row, column=1).value
            if not work_id or not str(work_id).strip():
                break
            
            normalized_id = self.normalize(work_id)
            
            if normalized_id not in target_index:
                self.warnings.append({'sheet': PROJECT_SUPPORT_SHEET, 'field': f'WorkID: {work_id}', 'reason': 'Not found in target'})
                self.ps_stats['not_found'] += 1
                continue
            
            tgt_row = target_index[normalized_id]
            
            if not self.dry_run:
                for col in range(1, PS_MAX_COL + 1):
                    source_cell = source_sheet.cell(row=src_row, column=col)
                    target_cell = target_sheet.cell(row=tgt_row, column=col)
                    if not self.is_formula(target_cell):
                        if not self.is_empty(source_cell.value) or self.is_empty(target_cell.value):
                            target_cell.value = source_cell.value
                            cells_copied += 1
            
            self.ps_stats['matched'] += 1
        
        return cells_copied
    
    def process_staff(self, source_wb, target_wb) -> int:
        """Process Staff sheet (p# and s# WorkIDs only)."""
        if STAFF_SHEET not in source_wb.sheetnames:
            return 0
        if STAFF_SHEET not in target_wb.sheetnames:
            return 0
        
        source_sheet = source_wb[STAFF_SHEET]
        target_sheet = target_wb[STAFF_SHEET]
        
        logger.info(f"Processing sheet '{STAFF_SHEET}' (p#/s# only)")
        
        # Collect valid source rows
        src_rows = []
        for row in range(STAFF_DATA_START_ROW, source_sheet.max_row + 1):
            work_id = source_sheet.cell(row=row, column=1).value
            if not work_id or not str(work_id).strip():
                break
            if self.is_valid_staff_workid(work_id):
                src_rows.append(row)
            else:
                logger.info(f"  Skipping row {row}: WorkID '{work_id}' (not p# or s#)")
                self.staff_stats['rows_skipped'] += 1
        
        src_count = len(src_rows)
        
        # Count target rows
        tgt_count = 0
        for row in range(STAFF_DATA_START_ROW, target_sheet.max_row + 1):
            if target_sheet.cell(row=row, column=1).value:
                tgt_count += 1
            else:
                break
        
        logger.info(f"  Source: {src_count} valid rows, Target: {tgt_count} rows")
        
        if self.dry_run:
            return 0
        
        cells_copied = 0
        
        # Copy valid rows
        for i, src_row in enumerate(src_rows):
            tgt_row = STAFF_DATA_START_ROW + i
            
            for col in STAFF_COPY_COLS:
                target_sheet.cell(row=tgt_row, column=col).value = source_sheet.cell(row=src_row, column=col).value
                cells_copied += 1
            
            # Update formula in column B
            target_sheet.cell(row=tgt_row, column=2).value = f'=VLOOKUP(A{tgt_row},tbl_projectNames,2,FALSE)'
        
        self.staff_stats['rows_copied'] = src_count
        
        # Clear extra rows
        if src_count < tgt_count:
            for i in range(tgt_count - src_count):
                row_to_clear = STAFF_DATA_START_ROW + src_count + i
                for col in range(1, target_sheet.max_column + 1):
                    target_sheet.cell(row=row_to_clear, column=col).value = None
            self.staff_stats['rows_deleted'] = tgt_count - src_count
        
        return cells_copied
    
    def add_logo_to_target(self, target_wb) -> bool:
        """Add company logo to target workbook at B15."""
        if not self.image_info:
            return False
        if LOGO_SHEET not in target_wb.sheetnames:
            return False
        
        target_sheet = target_wb[LOGO_SHEET]
        target_sheet._images = []
        target_sheet.row_dimensions[LOGO_ROW].height = LOGO_ROW_HEIGHT
        
        new_img = Image(self.image_info['path'])
        new_img.width = self.image_info['new_size'][0]
        new_img.height = self.image_info['new_size'][1]
        new_img.anchor = LOGO_CELL
        target_sheet.add_image(new_img)
        
        logger.info(f"Added logo at {LOGO_CELL}")
        return True
    
    def run(self) -> Dict[str, Any]:
        """Execute the CAS Plan setup."""
        logger.info("=" * 60)
        logger.info("CAS Plan Setup - Starting (v3.1.0)")
        logger.info("=" * 60)
        
        if not self.validate_files():
            return {'success': False, 'error': 'File validation failed'}
        
        self.image_info = self.extract_company_logo()
        
        try:
            source_wb = openpyxl.load_workbook(self.source_path, keep_vba=True)
            target_wb = openpyxl.load_workbook(self.target_path, keep_vba=True)
        except Exception as e:
            return {'success': False, 'error': str(e)}
        
        if not self.load_field_map(target_wb):
            source_wb.close()
            target_wb.close()
            return {'success': False, 'error': '_FieldMap not found'}
        
        total_cells = 0
        sheet_results = {}
        
        for sheet_name in FIELDMAP_SHEETS:
            cells = self.process_fieldmap_sheet(source_wb, target_wb, sheet_name)
            sheet_results[sheet_name] = cells
            total_cells += cells
        
        ps_cells = self.process_project_support(source_wb, target_wb)
        sheet_results[PROJECT_SUPPORT_SHEET] = ps_cells
        total_cells += ps_cells
        
        staff_cells = self.process_staff(source_wb, target_wb)
        sheet_results[STAFF_SHEET] = staff_cells
        total_cells += staff_cells
        
        images_copied = 0
        if not self.dry_run and self.image_info:
            if self.add_logo_to_target(target_wb):
                images_copied = 1
        
        if not self.dry_run:
            try:
                target_wb.save(self.target_path)
                source_wb.save(self.source_path)
            except Exception as e:
                source_wb.close()
                target_wb.close()
                return {'success': False, 'error': str(e)}
        
        source_wb.close()
        target_wb.close()
        
        return {
            'success': True,
            'total_cells_copied': total_cells,
            'images_copied': images_copied,
            'sheet_results': sheet_results,
            'ps_stats': self.ps_stats,
            'staff_stats': self.staff_stats,
            'preserved_count': len(self.preserved),
            'anomalies_count': len(self.anomalies)
        }
    
    def generate_report(self, results: Dict[str, Any], output_path: Optional[str] = None):
        """Generate report."""
        lines = [
            "=" * 70,
            "CAS PLAN SETUP REPORT (v3.1.0)",
            "=" * 70,
            f"Total Cells: {results.get('total_cells_copied', 0)}",
            "",
            "SHEET RESULTS:",
        ]
        for sheet, count in results.get('sheet_results', {}).items():
            lines.append(f"  {sheet}: {count}")
        
        ps = results.get('ps_stats', {})
        lines.extend([
            "",
            "PROJECT&SUPPORT:",
            f"  Matched: {ps.get('matched', 0)}, Columns: A-AB",
        ])
        
        staff = results.get('staff_stats', {})
        lines.extend([
            "",
            "STAFF:",
            f"  Rows: {staff.get('rows_copied', 0)}, Cleared: {staff.get('rows_deleted', 0)}, Skipped: {staff.get('rows_skipped', 0)}",
        ])
        
        report = "\n".join(lines)
        if output_path:
            with open(output_path, 'w') as f:
                f.write(report)
        return report


def main():
    parser = argparse.ArgumentParser(description='Setup CAS Plan (v3.1.0)')
    parser.add_argument('--source', '-s', required=True)
    parser.add_argument('--target', '-t', required=True)
    parser.add_argument('--dry-run', '-n', action='store_true')
    parser.add_argument('--backup', '-b', action='store_true')
    parser.add_argument('--report', '-r')
    
    args = parser.parse_args()
    setup = CASPlanSetup(args.source, args.target, args.dry_run)
    
    if args.backup and not args.dry_run:
        setup.create_backup(Path(args.source))
        setup.create_backup(Path(args.target))
    
    results = setup.run()
    print(setup.generate_report(results, args.report))
    exit(0 if results.get('success') else 1)


if __name__ == '__main__':
    main()
